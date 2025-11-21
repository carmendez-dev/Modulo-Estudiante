"""
Controlador para importar y exportar datos de estudiantes desde/hacia Excel
"""
from sqlalchemy.orm import Session
from fastapi import HTTPException, status, UploadFile
from app.models.estudiante_model import Estudiante
from typing import List, BinaryIO
import pandas as pd
from io import BytesIO
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

class ExcelController:
    """
    Controlador para operaciones de importación/exportación Excel
    """
    
    @staticmethod
    def exportar_estudiantes(db: Session) -> BytesIO:
        """
        Exportar todos los estudiantes a un archivo Excel
        
        Args:
            db: Sesión de base de datos
            
        Returns:
            BytesIO con el archivo Excel
        """
        # Obtener todos los estudiantes
        estudiantes = db.query(Estudiante).all()
        
        if not estudiantes:
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail="No hay estudiantes para exportar"
            )
        
        # Crear lista de diccionarios con los datos
        datos = []
        for est in estudiantes:
            datos.append({
                'ID': est.id_estudiante,
                'CI': est.ci,
                'Nombres': est.nombres,
                'Apellido Paterno': est.apellido_paterno,
                'Apellido Materno': est.apellido_materno,
                'Fecha Nacimiento': est.fecha_nacimiento.strftime('%Y-%m-%d') if est.fecha_nacimiento else '',
                'Dirección': est.direccion or '',
                'Estado': est.estado_estudiante,
                'Nombre Padre': est.nombre_padre or '',
                'Apellido Paterno Padre': est.apellido_paterno_padre or '',
                'Apellido Materno Padre': est.apellido_materno_padre or '',
                'Teléfono Padre': est.telefono_padre or '',
                'Nombre Madre': est.nombre_madre or '',
                'Apellido Paterno Madre': est.apellido_paterno_madre or '',
                'Apellido Materno Madre': est.apellido_materno_madre or '',
                'Teléfono Madre': est.telefono_madre or ''
            })
        
        # Crear DataFrame
        df = pd.DataFrame(datos)
        
        # Crear archivo Excel en memoria con formato
        output = BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name='Estudiantes')
            
            # Obtener el workbook y worksheet para aplicar estilos
            workbook = writer.book
            worksheet = writer.sheets['Estudiantes']
            
            # Estilos para el encabezado
            header_fill = PatternFill(start_color='27C5DA', end_color='27C5DA', fill_type='solid')
            header_font = Font(bold=True, color='FFFFFF', size=12)
            header_alignment = Alignment(horizontal='center', vertical='center')
            
            # Aplicar estilos al encabezado
            for cell in worksheet[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = header_alignment
            
            # Ajustar ancho de columnas
            for column in worksheet.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                worksheet.column_dimensions[column_letter].width = adjusted_width
            
            # Agregar bordes
            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            for row in worksheet.iter_rows(min_row=1, max_row=len(datos)+1, min_col=1, max_col=len(df.columns)):
                for cell in row:
                    cell.border = thin_border
        
        output.seek(0)
        return output
    
    @staticmethod
    def exportar_estudiante_por_id(db: Session, id_estudiante: int) -> BytesIO:
        """
        Exportar un estudiante específico a un archivo Excel
        
        Args:
            db: Sesión de base de datos
            id_estudiante: ID del estudiante a exportar
            
        Returns:
            BytesIO con el archivo Excel
        """
        # Obtener el estudiante
        estudiante = db.query(Estudiante).filter(
            Estudiante.id_estudiante == id_estudiante
        ).first()
        
        if not estudiante:
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail=f"Estudiante con ID {id_estudiante} no encontrado"
            )
        
        # Crear diccionario con los datos del estudiante
        datos = [{
            'ID': estudiante.id_estudiante,
            'CI': estudiante.ci,
            'Nombres': estudiante.nombres,
            'Apellido Paterno': estudiante.apellido_paterno,
            'Apellido Materno': estudiante.apellido_materno,
            'Fecha Nacimiento': estudiante.fecha_nacimiento.strftime('%Y-%m-%d') if estudiante.fecha_nacimiento else '',
            'Dirección': estudiante.direccion or '',
            'Estado': estudiante.estado_estudiante,
            'Nombre Padre': estudiante.nombre_padre or '',
            'Apellido Paterno Padre': estudiante.apellido_paterno_padre or '',
            'Apellido Materno Padre': estudiante.apellido_materno_padre or '',
            'Teléfono Padre': estudiante.telefono_padre or '',
            'Nombre Madre': estudiante.nombre_madre or '',
            'Apellido Paterno Madre': estudiante.apellido_paterno_madre or '',
            'Apellido Materno Madre': estudiante.apellido_materno_madre or '',
            'Teléfono Madre': estudiante.telefono_madre or ''
        }]
        
        # Crear DataFrame
        df = pd.DataFrame(datos)
        
        # Crear archivo Excel en memoria con formato
        output = BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name='Estudiante')
            
            # Obtener el workbook y worksheet para aplicar estilos
            workbook = writer.book
            worksheet = writer.sheets['Estudiante']
            
            # Estilos para el encabezado
            header_fill = PatternFill(start_color='27C5DA', end_color='27C5DA', fill_type='solid')
            header_font = Font(bold=True, color='FFFFFF', size=12)
            header_alignment = Alignment(horizontal='center', vertical='center')
            
            # Aplicar estilos al encabezado
            for cell in worksheet[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = header_alignment
            
            # Ajustar ancho de columnas
            for column in worksheet.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                worksheet.column_dimensions[column_letter].width = adjusted_width
            
            # Agregar bordes
            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            for row in worksheet.iter_rows(min_row=1, max_row=2, min_col=1, max_col=len(df.columns)):
                for cell in row:
                    cell.border = thin_border
            
            # Agregar información de cursos en una segunda hoja si tiene cursos
            if estudiante.cursos:
                cursos_data = []
                for curso in estudiante.cursos:
                    cursos_data.append({
                        'ID Curso': curso.id_curso,
                        'Nombre Curso': curso.nombre_curso,
                        'Nivel': curso.nivel,
                        'Gestión': curso.gestion
                    })
                
                df_cursos = pd.DataFrame(cursos_data)
                df_cursos.to_excel(writer, index=False, sheet_name='Cursos')
                
                worksheet_cursos = writer.sheets['Cursos']
                
                # Aplicar estilos al encabezado de cursos
                for cell in worksheet_cursos[1]:
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = header_alignment
                
                # Ajustar ancho de columnas
                for column in worksheet_cursos.columns:
                    max_length = 0
                    column_letter = column[0].column_letter
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = min(max_length + 2, 30)
                    worksheet_cursos.column_dimensions[column_letter].width = adjusted_width
                
                # Agregar bordes
                for row in worksheet_cursos.iter_rows(min_row=1, max_row=len(cursos_data)+1, min_col=1, max_col=len(df_cursos.columns)):
                    for cell in row:
                        cell.border = thin_border
        
        output.seek(0)
        return output
    
    @staticmethod
    def importar_estudiantes(db: Session, file: UploadFile) -> dict:
        """
        Importar estudiantes desde un archivo Excel
        
        Args:
            db: Sesión de base de datos
            file: Archivo Excel subido
            
        Returns:
            Diccionario con resultado de la importación
        """
        # Validar extensión del archivo
        if not file.filename.endswith(('.xlsx', '.xls')):
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="El archivo debe ser un Excel (.xlsx o .xls)"
            )
        
        try:
            # Leer el archivo Excel
            contents = file.file.read()
            df = pd.read_excel(BytesIO(contents))
            
            # Validar columnas requeridas
            columnas_requeridas = ['CI', 'Nombres', 'Apellido Paterno', 'Apellido Materno']
            columnas_faltantes = [col for col in columnas_requeridas if col not in df.columns]
            
            if columnas_faltantes:
                raise HTTPException(
                    status_code=status.HTTP_400_BAD_REQUEST,
                    detail=f"Faltan columnas requeridas: {', '.join(columnas_faltantes)}"
                )
            
            # Reemplazar NaN con None
            df = df.where(pd.notna(df), None)
            
            estudiantes_creados = 0
            estudiantes_actualizados = 0
            errores = []
            
            for index, row in df.iterrows():
                try:
                    # Verificar si el estudiante ya existe por CI
                    estudiante_existente = None
                    if row.get('CI'):
                        estudiante_existente = db.query(Estudiante).filter(
                            Estudiante.ci == str(row['CI'])
                        ).first()
                    
                    # Preparar datos del estudiante
                    datos_estudiante = {
                        'ci': str(row['CI']) if row.get('CI') else None,
                        'nombres': row['Nombres'],
                        'apellido_paterno': row['Apellido Paterno'],
                        'apellido_materno': row['Apellido Materno'],
                        'fecha_nacimiento': pd.to_datetime(row['Fecha Nacimiento']).date() if row.get('Fecha Nacimiento') and pd.notna(row['Fecha Nacimiento']) else None,
                        'direccion': row.get('Dirección'),
                        'estado_estudiante': row.get('Estado', 'Activo'),
                        'nombre_padre': row.get('Nombre Padre'),
                        'apellido_paterno_padre': row.get('Apellido Paterno Padre'),
                        'apellido_materno_padre': row.get('Apellido Materno Padre'),
                        'telefono_padre': row.get('Teléfono Padre'),
                        'nombre_madre': row.get('Nombre Madre'),
                        'apellido_paterno_madre': row.get('Apellido Paterno Madre'),
                        'apellido_materno_madre': row.get('Apellido Materno Madre'),
                        'telefono_madre': row.get('Teléfono Madre')
                    }
                    
                    if estudiante_existente:
                        # Actualizar estudiante existente
                        for campo, valor in datos_estudiante.items():
                            if valor is not None:
                                setattr(estudiante_existente, campo, valor)
                        estudiantes_actualizados += 1
                    else:
                        # Crear nuevo estudiante
                        nuevo_estudiante = Estudiante(**datos_estudiante)
                        db.add(nuevo_estudiante)
                        estudiantes_creados += 1
                    
                except Exception as e:
                    errores.append(f"Fila {index + 2}: {str(e)}")
            
            # Confirmar cambios
            db.commit()
            
            return {
                "mensaje": "Importación completada",
                "estudiantes_creados": estudiantes_creados,
                "estudiantes_actualizados": estudiantes_actualizados,
                "total_procesados": estudiantes_creados + estudiantes_actualizados,
                "errores": errores if errores else None
            }
            
        except HTTPException:
            raise
        except Exception as e:
            db.rollback()
            raise HTTPException(
                status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
                detail=f"Error al procesar el archivo Excel: {str(e)}"
            )
    
    @staticmethod
    def descargar_plantilla() -> BytesIO:
        """
        Generar una plantilla Excel vacía para importar estudiantes
        
        Returns:
            BytesIO con la plantilla Excel
        """
        # Crear DataFrame con columnas y datos de ejemplo
        datos_ejemplo = [{
            'CI': '12345678',
            'Nombres': 'Juan Carlos',
            'Apellido Paterno': 'Pérez',
            'Apellido Materno': 'García',
            'Fecha Nacimiento': '2010-05-15',
            'Dirección': 'Av. Principal #123',
            'Estado': 'Activo',
            'Nombre Padre': 'Carlos',
            'Apellido Paterno Padre': 'Pérez',
            'Apellido Materno Padre': 'López',
            'Teléfono Padre': '70000001',
            'Nombre Madre': 'María',
            'Apellido Paterno Madre': 'García',
            'Apellido Materno Madre': 'Rojas',
            'Teléfono Madre': '70000002'
        }]
        
        df = pd.DataFrame(datos_ejemplo)
        
        # Crear archivo Excel en memoria con formato
        output = BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name='Estudiantes')
            
            workbook = writer.book
            worksheet = writer.sheets['Estudiantes']
            
            # Estilos para el encabezado
            header_fill = PatternFill(start_color='3AC0B8', end_color='3AC0B8', fill_type='solid')
            header_font = Font(bold=True, color='FFFFFF', size=12)
            header_alignment = Alignment(horizontal='center', vertical='center')
            
            # Aplicar estilos al encabezado
            for cell in worksheet[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = header_alignment
            
            # Ajustar ancho de columnas
            for column in worksheet.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                worksheet.column_dimensions[column_letter].width = adjusted_width
            
            # Agregar instrucciones en una hoja separada
            instrucciones = [
                ['INSTRUCCIONES PARA IMPORTAR ESTUDIANTES'],
                [''],
                ['Columnas Requeridas (obligatorias):'],
                ['- CI: Cédula de Identidad del estudiante'],
                ['- Nombres: Nombres del estudiante'],
                ['- Apellido Paterno: Apellido paterno del estudiante'],
                ['- Apellido Materno: Apellido materno del estudiante'],
                [''],
                ['Columnas Opcionales:'],
                ['- Fecha Nacimiento: Formato YYYY-MM-DD (ej: 2010-05-15)'],
                ['- Dirección: Dirección del estudiante'],
                ['- Estado: Activo, Retirado o Abandono (por defecto: Activo)'],
                ['- Nombre Padre: Nombre del padre'],
                ['- Apellido Paterno Padre: Apellido paterno del padre'],
                ['- Apellido Materno Padre: Apellido materno del padre'],
                ['- Teléfono Padre: Teléfono del padre'],
                ['- Nombre Madre: Nombre de la madre'],
                ['- Apellido Paterno Madre: Apellido paterno de la madre'],
                ['- Apellido Materno Madre: Apellido materno de la madre'],
                ['- Teléfono Madre: Teléfono de la madre'],
                [''],
                ['Notas Importantes:'],
                ['1. Si el CI ya existe, se actualizará el estudiante'],
                ['2. Si el CI no existe, se creará un nuevo estudiante'],
                ['3. La primera fila contiene datos de ejemplo, puede eliminarla'],
                ['4. No modifique los nombres de las columnas'],
                ['5. Guarde el archivo como .xlsx antes de importar']
            ]
            
            df_instrucciones = pd.DataFrame(instrucciones)
            df_instrucciones.to_excel(writer, index=False, header=False, sheet_name='Instrucciones')
            
            worksheet_inst = writer.sheets['Instrucciones']
            
            # Estilo para el título
            worksheet_inst['A1'].font = Font(bold=True, size=14, color='0B2E50')
            worksheet_inst['A1'].fill = PatternFill(start_color='27C5DA', end_color='27C5DA', fill_type='solid')
            
            # Ajustar ancho de columna
            worksheet_inst.column_dimensions['A'].width = 80
        
        output.seek(0)
        return output
